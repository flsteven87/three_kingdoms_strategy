"""One-off script to export copper coordinates Excel to SQL."""

import re
from pathlib import Path

from openpyxl import load_workbook

FILE = Path(r"C:\Users\Jason\Downloads\PK23霸王討逆-銅礦坐標.xlsx")
SEASON_TAG = "PK21"
OUTPUT = Path(r"C:\Users\Jason\Downloads\pk21_copper_coordinates.sql")

wb = load_workbook(FILE, read_only=True, data_only=True)
ws = wb.active
if ws is None:
    raise ValueError("No active worksheet")

rows: list[tuple[int, int, int, str, str]] = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or all(c is None for c in row):
        continue
    if len(row) < 5:
        continue
    _, lv, coord, county, district = row[0], row[1], row[2], row[3], row[4]
    if not all([lv, coord, county, district]):
        continue
    m = re.match(r"\(?\s*(\d+)\s*,\s*(\d+)\s*\)?", str(coord).strip())
    if not m:
        continue
    rows.append((
        int(m.group(1)),
        int(m.group(2)),
        int(lv),
        str(county).strip(),
        str(district).strip(),
    ))
wb.close()

lines = []
for x, y, lv, county, district in rows:
    c = county.replace("'", "''")
    d = district.replace("'", "''")
    lines.append(f"  ('{SEASON_TAG}', {x}, {y}, {lv}, '{c}', '{d}')")

sql = (
    "INSERT INTO copper_mine_coordinates (game_season_tag, coord_x, coord_y, level, county, district)\nVALUES\n"
    + ",\n".join(lines)
    + "\nON CONFLICT (game_season_tag, coord_x, coord_y)\n"
    + "DO UPDATE SET level = EXCLUDED.level, county = EXCLUDED.county, district = EXCLUDED.district;\n"
)

OUTPUT.write_text(sql, encoding="utf-8")
print(f"Wrote {len(rows)} rows to {OUTPUT}")
