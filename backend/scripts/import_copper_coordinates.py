"""
Import copper mine coordinates from Excel into copper_mine_coordinates table.

Usage:
    uv run python backend/scripts/import_copper_coordinates.py \
        --file "PK23霸王討逆-銅礦坐標.xlsx" \
        --season-tag PK23

Excel columns (order):
    1. Resource name (資源名稱) — e.g. "銅礦"
    2. Level (等級) — 8, 9, or 10
    3. Coordinate (坐標) — format "(x, y)"
    4. County (郡)
    5. District (縣)
"""

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def parse_coordinate(raw: str) -> tuple[int, int]:
    """
    Parse coordinate string in "(x, y)" format.

    Args:
        raw: Coordinate string, e.g. "(123, 456)" or "(123,456)"

    Returns:
        Tuple of (x, y) integers

    Raises:
        ValueError: If format is invalid
    """
    match = re.match(r"\(?\s*(\d+)\s*,\s*(\d+)\s*\)?", str(raw).strip())
    if not match:
        raise ValueError(f"Invalid coordinate format: {raw!r}")
    return int(match.group(1)), int(match.group(2))


def read_excel(file_path: Path) -> list[dict]:
    """
    Read copper mine data from Excel file.

    Expected column order: resource_name, level, coordinate, county, district
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no active worksheet")

    rows: list[dict] = []
    skipped = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not row or all(cell is None for cell in row):
            continue

        if len(row) < 5:
            print(f"  Warning: Row {row_num} has fewer than 5 columns, skipping")
            skipped += 1
            continue

        _, level_raw, coord_raw, county, district = row[0], row[1], row[2], row[3], row[4]

        # Skip rows with missing required fields
        if not all([level_raw, coord_raw, county, district]):
            print(f"  Warning: Row {row_num} has missing fields, skipping: {row[:5]}")
            skipped += 1
            continue

        try:
            level = int(level_raw)
        except (ValueError, TypeError):
            print(f"  Warning: Row {row_num} has invalid level {level_raw!r}, skipping")
            skipped += 1
            continue

        if level < 1 or level > 10:
            print(f"  Warning: Row {row_num} has out-of-range level {level}, skipping")
            skipped += 1
            continue

        try:
            coord_x, coord_y = parse_coordinate(str(coord_raw))
        except ValueError as e:
            print(f"  Warning: Row {row_num}: {e}, skipping")
            skipped += 1
            continue

        rows.append({
            "county": str(county).strip(),
            "district": str(district).strip(),
            "coord_x": coord_x,
            "coord_y": coord_y,
            "level": level,
        })

    wb.close()

    if skipped:
        print(f"  Skipped {skipped} rows due to errors")

    return rows


def import_to_supabase(rows: list[dict], season_tag: str, *, dry_run: bool = False) -> int:
    """
    Upsert rows into copper_mine_coordinates via Supabase client.

    Args:
        rows: Parsed coordinate data
        season_tag: Game season identifier (e.g. "PK23")
        dry_run: If True, only print what would be inserted

    Returns:
        Number of rows upserted
    """
    if dry_run:
        print(f"\n[DRY RUN] Would upsert {len(rows)} rows with game_season_tag={season_tag!r}")
        for i, row in enumerate(rows[:5]):
            print(f"  {i+1}. ({row['coord_x']}, {row['coord_y']}) Lv.{row['level']} — {row['county']} {row['district']}")
        if len(rows) > 5:
            print(f"  ... and {len(rows) - 5} more rows")
        return 0

    # Import here to avoid requiring env vars during --dry-run
    import os

    from supabase import create_client

    supabase_url = os.environ.get("SUPABASE_URL")
    supabase_service_key = os.environ.get("SUPABASE_SERVICE_KEY")

    if not supabase_url or not supabase_service_key:
        print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY environment variables are required")
        sys.exit(1)

    client = create_client(supabase_url, supabase_service_key)

    # Prepare records with game_season_tag
    records = [
        {
            "game_season_tag": season_tag,
            "county": row["county"],
            "district": row["district"],
            "coord_x": row["coord_x"],
            "coord_y": row["coord_y"],
            "level": row["level"],
        }
        for row in rows
    ]

    # Upsert in batches of 500
    batch_size = 500
    total_upserted = 0

    for i in range(0, len(records), batch_size):
        batch = records[i : i + batch_size]
        result = (
            client.table("copper_mine_coordinates")
            .upsert(batch, on_conflict="game_season_tag,coord_x,coord_y")
            .execute()
        )
        total_upserted += len(result.data)
        print(f"  Upserted batch {i // batch_size + 1}: {len(result.data)} rows")

    return total_upserted


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import copper mine coordinates from Excel into Supabase"
    )
    parser.add_argument(
        "--file", required=True, type=Path, help="Path to the Excel (.xlsx) file"
    )
    parser.add_argument(
        "--season-tag", required=True, help="Game season tag (e.g. PK23)"
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="Parse and validate only, do not write to DB"
    )

    args = parser.parse_args()

    if not args.file.exists():
        print(f"Error: File not found: {args.file}")
        sys.exit(1)

    print(f"Reading {args.file}...")
    rows = read_excel(args.file)
    print(f"Parsed {len(rows)} valid rows")

    if not rows:
        print("No valid rows found. Exiting.")
        sys.exit(1)

    # Summary
    levels = {}
    counties = set()
    for row in rows:
        levels[row["level"]] = levels.get(row["level"], 0) + 1
        counties.add(row["county"])

    print(f"\nSummary for {args.season_tag}:")
    for lvl in sorted(levels):
        print(f"  Level {lvl}: {levels[lvl]} mines")
    print(f"  Counties: {len(counties)} ({', '.join(sorted(counties)[:5])}{'...' if len(counties) > 5 else ''})")

    count = import_to_supabase(rows, args.season_tag, dry_run=args.dry_run)

    if not args.dry_run:
        print(f"\nDone! Upserted {count} rows into copper_mine_coordinates.")


if __name__ == "__main__":
    main()
